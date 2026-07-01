"""
IMPL-20260630-03: Generador XLSX para reportes masivos.
ARCH-20260623-01: Modulo de Reportes Masivos.

Replica la estructura 3 hojas del formato operativo:
  - CONCENTRADO   (1 fila por trabajador)
  - LABORATORIOS  (1 fila por trabajador, paneles BH/QS6/EGO/TOXICO)
  - GRAFICAS      (agregados/secciones)
"""
from __future__ import annotations

import os
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.services.reports.conteos import (
    calcular_escoliosis_distribucion,
    calcular_espirometria_distribucion,
    calcular_hbc_por_rango,
    calcular_qs6_niveles,
    calcular_trauma_acustico_por_area,
)


HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _str_or_na(v: Any) -> str:
    if v is None:
        return "N/A"
    if isinstance(v, str):
        return v if v.strip() else "N/A"
    return str(v)


def _fila_concentrado(w: Dict[str, Any]) -> List[Any]:
    a = w.get("audiometria") or {}
    e = w.get("espirometria") or {}
    rxc = w.get("rxColumna") or {}
    rxt = w.get("rxTorax") or {}
    ecg = w.get("ecg") or {}
    c = w.get("campimetria") or {}
    return [
        _str_or_na(w.get("folio")),
        _str_or_na(w.get("nombre")),
        _str_or_na(w.get("sexo")),
        _str_or_na(w.get("area")),
        _str_or_na(w.get("antiguedad")),
        _str_or_na(c.get("agudezaVisual")),
        _str_or_na(c.get("camposVisuales")),
        _str_or_na(c.get("discriminacionColor")),
        _str_or_na(a.get("dx")),
        _str_or_na(a.get("oidoDerecho")),
        _str_or_na(a.get("oidoIzquierdo")),
        _str_or_na(a.get("hbc")),
        _str_or_na(e.get("patron")),
        _str_or_na(e.get("fvc")),
        _str_or_na(e.get("tabaquismo")),
        _str_or_na(ecg.get("impresion")),
        _str_or_na(rxc.get("valoracionPostural")),
        _str_or_na(rxc.get("escoliosis")),
        _str_or_na(rxc.get("lordosis")),
        _str_or_na(rxc.get("basculacion")),
        _str_or_na(rxc.get("impresion")),
        _str_or_na(rxt.get("impresion")),
    ]


CONCENTRADO_HEADERS = [
    "FOLIO",
    "NOMBRE",
    "SEXO",
    "AREA/ PUESTO",
    "ANTIGÜEDAD",
    "AGUDEZA VISUAL",
    "CAMPOS VISUALES",
    "DISCRIMINACION DEL COLOR",
    "DX",
    "OIDO DERECHO",
    "OIDO IZQUIERDO",
    "% HBC",
    "ESPIROMETRIA",
    "FVC",
    "TABAQUISMO",
    "ELECTROCARDIOGRAMA",
    "VALORACION POSTURAL",
    "GRADO DE ESCOLIOSIS(°)",
    "GRADO DE LORDOSIS(°)",
    "BASCULACIÓN PELVICA (cms)",
    "IMPRESIÓN DIAGNOSTICA COLUMNA",
    "IMPRESIÓN DIAGNOSTICA TORAX",
]


def _fila_laboratorio(w: Dict[str, Any]) -> List[Any]:
    lab = w.get("laboratorio") or {}
    bh = lab.get("bh") or {}
    qs6 = lab.get("qs6") or {}
    ego = lab.get("ego") or {}
    tox = lab.get("toxico") or {}
    return [
        _str_or_na(w.get("folio")),
        _str_or_na(lab.get("folioLab")),
        _str_or_na(w.get("nombre")),
        _str_or_na(w.get("sexo")),
        _str_or_na(lab.get("edad")),
        _str_or_na(bh.get("hb")),
        _str_or_na(bh.get("mchb")),
        _str_or_na(bh.get("chgm")),
        _str_or_na(bh.get("leu")),
        _str_or_na(bh.get("pla")),
        _str_or_na(qs6.get("gluc")),
        _str_or_na(qs6.get("bun")),
        _str_or_na(qs6.get("urea")),
        _str_or_na(qs6.get("creat")),
        _str_or_na(qs6.get("au")),
        _str_or_na(qs6.get("col")),
        _str_or_na(qs6.get("trig")),
        _str_or_na(ego.get("glc")),
        _str_or_na(ego.get("prot")),
        _str_or_na(ego.get("blo")),
        _str_or_na(ego.get("bac")),
        _str_or_na(ego.get("cristales")),
        _str_or_na(tox.get("anfeta")),
        _str_or_na(tox.get("coca")),
        _str_or_na(tox.get("marihua")),
        _str_or_na(tox.get("opiac")),
        _str_or_na(tox.get("metanf")),
    ]


LABORATORIOS_HEADERS = [
    "Folio", "Folio Lab", "Nombre", "SEXO", "Edad",
    "BH Hb", "BH MCHb", "BH CHGM", "BH LEU", "BH PLA",
    "GLUC", "BUN", "UREA", "CREAT", "AU", "COL", "TRIG",
    "EGO-GLC", "EGO-PROT", "EGO-BLO", "EGO BAC", "CRISTALES EGO",
    "ANFETA", "COCA", "MARIHUA", "OPIAC", "METANF",
]


def _estilo_header(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def generar_xlsx(project: Dict[str, Any], output_path: str) -> str:
    """
    Genera el archivo XLSX de 3 hojas en `output_path`.
    Retorna la ruta absoluta generada.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()

    # Hoja 1: CONCENTRADO
    ws1 = wb.active
    ws1.title = "CONCENTRADO"
    ws1.append(CONCENTRADO_HEADERS)
    for w in project.get("trabajadores", []):
        ws1.append(_fila_concentrado(w))
    _estilo_header(ws1, len(CONCENTRADO_HEADERS))

    # Hoja 2: LABORATORIOS
    ws2 = wb.create_sheet("LABORATORIOS")
    ws2.append(LABORATORIOS_HEADERS)
    for w in project.get("trabajadores", []):
        ws2.append(_fila_laboratorio(w))
    _estilo_header(ws2, len(LABORATORIOS_HEADERS))

    # Hoja 3: GRAFICAS (agregados)
    ws3 = wb.create_sheet("GRAFICAS")
    trauma = calcular_trauma_acustico_por_area(project)
    hbc = calcular_hbc_por_rango(project)
    espiro = calcular_espirometria_distribucion(project)
    escolio = calcular_escoliosis_distribucion(project)
    qs6 = calcular_qs6_niveles(project)

    rows: List[List[Any]] = []
    rows.append(["TRAUMA ACUSTICO POR AREA"])
    rows.append(["AREA", "TRABAJADORES"])
    for t in trauma:
        rows.append([t["area"], t["conteo"]])
    rows.append([])
    rows.append(["AUDIOMETRIAS (% HBC)"])
    rows.append(["RANGO", "TRABAJADORES"])
    rows.append(["Normal (<10%)", hbc["normal"]])
    rows.append(["Alto (10-19%)", hbc["alto"]])
    rows.append(["Muy Alto (>=20%)", hbc["muyAlto"]])
    rows.append([])
    rows.append(["ESPIROMETRIAS (PATRON)"])
    rows.append(["PATRON", "TRABAJADORES"])
    for e in espiro:
        rows.append([e["patron"], e["conteo"]])
    rows.append([])
    rows.append(["COLUMNA (ESCOLIOSIS - COBB)"])
    rows.append(["GRADO", "TRABAJADORES"])
    rows.append(["NORMAL (<5°)", escolio["normal"]])
    rows.append(["LEVE (5-9°)", escolio["leve"]])
    rows.append(["MODERADA (10-19°)", escolio["moderada"]])
    rows.append(["GRAVE (>=20°)", escolio["grave"]])
    rows.append([])
    rows.append(["QS6 - GLUCOSA"])
    rows.append(["RANGO", "TRABAJADORES"])
    rows.append(["Normal (<100 mg/dL)", qs6["glucosa"]["normal"]])
    rows.append(["Alta (>=100 mg/dL)", qs6["glucosa"]["alta"]])
    rows.append([])
    rows.append(["QS6 - COLESTEROL"])
    rows.append(["RANGO", "TRABAJADORES"])
    rows.append(["Normal (<200 mg/dL)", qs6["colesterol"]["normal"]])
    rows.append(["Limite (200-239 mg/dL)", qs6["colesterol"]["limite"]])
    rows.append(["Alto (>=240 mg/dL)", qs6["colesterol"]["alto"]])
    rows.append([])
    rows.append(["QS6 - TRIGLICERIDOS"])
    rows.append(["RANGO", "TRABAJADORES"])
    rows.append(["Normal (<150 mg/dL)", qs6["trigliceridos"]["normal"]])
    rows.append(["Limite (150-199 mg/dL)", qs6["trigliceridos"]["limite"]])
    rows.append(["Alto (>=200 mg/dL)", qs6["trigliceridos"]["alto"]])

    for row in rows:
        ws3.append(row)

    wb.save(output_path)
    return output_path