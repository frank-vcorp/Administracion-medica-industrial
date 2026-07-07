"""
Tests pytest para el modulo de Reportes Masivos.
IMPL-20260701-04: Fase 4 EBOOK — actualizacion de orquestador + API.
IMPL-20260630-03: ARCH-20260623-01 (tests originales).

Cubre:
  1. test_create_project_report_returns_pending
  2. test_xlsx_writer_creates_3_sheets
  3. test_pdf_writer_creates_portada_and_concentrado  (legacy directo)
  4. test_conteos_calculation
  5. test_report_status_transitions_pending_to_ready
  6. test_report_failed_with_error_message
  + IMPL-20260701-04:
  7. test_generar_reporte_masivo_ebook     (nuevo)
  8. test_generar_reporte_masivo_both_ebook (actualizado)
  9. test_api_rejects_legacy_pdf_format    (nuevo)

Los tests de endpoints usan mocks del Prisma client para no depender
de una DB real. Los generadores se prueban con un snapshot construido
a mano (mismo shape que project_to_snapshot).
"""
import asyncio
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

# Permitir imports del paquete app.*
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.reports import conteos, pdf_writer, xlsx_writer  # noqa: E402
from app.services.reports.massive_report import (  # noqa: E402
    generar_reporte_masivo,
    run_generation_job,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def tmp_uploads(tmp_path, monkeypatch):
    """Redirige UPLOAD_DIR a un tmp_path y devuelve el path."""
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir()
    monkeypatch.setenv("UPLOAD_DIR", str(upload_dir))
    return upload_dir


def _make_worker(
    folio: str = "168058",
    nombre: str = "JUAN PEREZ",
    sexo: str = "MASCULINO",
    area: str = "SOLDADURA",
    *,
    completo: bool = True,
    hbc: float = 5.0,
    patron: str = "NORMAL",
    escoliosis: float = 3.0,
) -> dict:
    """Crea un dict con shape de DemoWorker para tests."""
    if completo:
        return {
            "folio": folio,
            "nombre": nombre,
            "sexo": sexo,
            "area": area,
            "antiguedad": "3 A",
            "campimetria": {
                "agudezaVisual": "NORMAL",
                "camposVisuales": "NORMAL",
                "discriminacionColor": "NORMAL",
            },
            "audiometria": {
                "dx": "NORMAL",
                "oidoDerecho": "Audicion Normal",
                "oidoIzquierdo": "Audicion Normal",
                "hbc": hbc,
            },
            "espirometria": {
                "patron": patron,
                "fvc": 0.93,
                "tabaquismo": "NEGADO",
            },
            "rxColumna": {
                "escoliosis": escoliosis,
                "lordosis": 36,
                "basculacion": 0,
                "valoracionPostural": "NORMAL",
                "impresion": "SIN HALLAZGOS",
            },
            "rxTorax": {"impresion": "TORAX NORMAL"},
            "ecg": {"impresion": "SIN ALTERACIONES"},
            "laboratorio": {
                "folioLab": "260417010016",
                "edad": "28 A",
                "bh": {"hb": 15.0, "mchb": 30.0, "chgm": 33.0, "leu": 7.0, "pla": 220.0},
                "qs6": {"gluc": 90.0, "bun": 15.0, "urea": 35.0, "creat": 1.0, "au": 5.0, "col": 180.0, "trig": 120.0},
                "ego": {"glc": 0, "prot": 0, "blo": 0, "bac": "AUSENTES", "cristales": "URATOS ESCASOS"},
                "toxico": {"anfeta": "NEGATIVO", "coca": "NEGATIVO", "marihua": "NEGATIVO", "opiac": "NEGATIVO", "metanf": "NEGATIVO"},
            },
        }
    # Trabajador con todos los estudios en N/A.
    return {
        "folio": folio,
        "nombre": nombre,
        "sexo": sexo,
        "area": area,
        "antiguedad": "1 A",
        "campimetria": {"agudezaVisual": "N/A", "camposVisuales": "N/A", "discriminacionColor": "N/A"},
        "audiometria": {"dx": "N/A", "oidoDerecho": "N/A", "oidoIzquierdo": "N/A", "hbc": None},
        "espirometria": {"patron": "N/A", "fvc": None, "tabaquismo": "N/A"},
        "rxColumna": {"escoliosis": None, "lordosis": None, "basculacion": None, "valoracionPostural": "N/A", "impresion": "N/A"},
        "rxTorax": {"impresion": "N/A"},
        "ecg": {"impresion": "N/A"},
        "laboratorio": {
            "folioLab": "", "edad": "",
            "bh": {"hb": None, "mchb": None, "chgm": None, "leu": None, "pla": None},
            "qs6": {"gluc": None, "bun": None, "urea": None, "creat": None, "au": None, "col": None, "trig": None},
            "ego": {"glc": None, "prot": None, "blo": None, "bac": "N/A", "cristales": "N/A"},
            "toxico": {"anfeta": "N/A", "coca": "N/A", "marihua": "N/A", "opiac": "N/A", "metanf": "N/A"},
        },
    }


@pytest.fixture
def sample_project() -> dict:
    """Proyecto de prueba con 3 trabajadores (1 completo, 1 parcial, 1 sin estudios)."""
    w1 = _make_worker(folio="1", nombre="AAA", completo=True)
    # w2: parcial — algunos campos N/A (ecg, rxTorax), resto con datos; hbc=18 (alto).
    w2 = _make_worker(folio="2", nombre="BBB", completo=True, hbc=18.0)
    w2["ecg"] = {"impresion": "N/A"}
    w2["rxTorax"] = {"impresion": "N/A"}

    w3 = _make_worker(folio="3", nombre="CCC", completo=False)
    # w3 ya es completo=False, le pongo area distinta.
    w3["area"] = "ALMACEN"

    return {
        "id": "proj-test",
        "empresa": "TEST S.A.",
        "empresaLegal": "TEST SOCIEDAD ANONIMA",
        "fecha": "2026-06-30",
        "trabajadores": [w1, w2, w3],
    }


# ---------------------------------------------------------------------------
# Tests de conteos
# ---------------------------------------------------------------------------


def test_conteos_calculation(sample_project):
    """Verifica el calculo de totales, completos, parciales y sin estudios."""
    result = conteos.calcular_conteos(sample_project)
    assert result["total"] == 3
    assert result["completos"] == 1   # solo w1
    assert result["sinEstudios"] == 1  # solo w3
    assert result["parciales"] == 1   # w2


def test_distribuciones_y_hbc(sample_project):
    """Verifica distribuciones por edad/sexo y rangos HBC."""
    dist = conteos.calcular_distribuciones(sample_project)
    assert dist["masculino"] == 3
    # w1=28A, w2=28A -> 18-30; w3 sin edad -> None (ignorado)
    assert dist["edad18a30"] == 2

    hbc = conteos.calcular_hbc_por_rango(sample_project)
    # w1=5 (normal), w2=18 (alto), w3=None (ignorado)
    assert hbc["normal"] == 1
    assert hbc["alto"] == 1
    assert hbc["muyAlto"] == 0


def test_escoliosis_y_qs6(sample_project):
    """Verifica distribuciones de escoliosis y qs6."""
    escolio = conteos.calcular_escoliosis_distribucion(sample_project)
    # w1=3 (normal), w2=3 (normal), w3=None (ignorado)
    assert escolio["normal"] == 2

    qs6 = conteos.calcular_qs6_niveles(sample_project)
    # w1 y w2 con datos (gluc=90 normal, col=180 normal, trig=120 normal); w3 con None -> ignorado.
    assert qs6["glucosa"]["normal"] == 2
    assert qs6["colesterol"]["normal"] == 2
    assert qs6["trigliceridos"]["normal"] == 2


# ---------------------------------------------------------------------------
# Tests XLSX
# ---------------------------------------------------------------------------


def test_xlsx_writer_creates_3_sheets(sample_project, tmp_path):
    """El XLSX debe contener exactamente las 3 hojas esperadas."""
    out = tmp_path / "concentrado.xlsx"
    xlsx_writer.generar_xlsx(sample_project, str(out))
    assert out.exists()
    wb = load_workbook(str(out))
    assert set(wb.sheetnames) == {"CONCENTRADO", "LABORATORIOS", "GRAFICAS"}
    # Cada hoja debe tener al menos 2 filas (header + 3 trabajadores).
    for name in wb.sheetnames:
        assert wb[name].max_row >= 2


def test_xlsx_writer_empty_project(tmp_path):
    """Proyecto sin trabajadores debe generar XLSX valido con headers solamente."""
    out = tmp_path / "empty.xlsx"
    xlsx_writer.generar_xlsx({"empresa": "EMPTY", "trabajadores": []}, str(out))
    assert out.exists()
    wb = load_workbook(str(out))
    assert "CONCENTRADO" in wb.sheetnames
    # Solo header
    assert wb["CONCENTRADO"].max_row == 1


# ---------------------------------------------------------------------------
# Tests PDF
# ---------------------------------------------------------------------------


def test_pdf_writer_creates_portada_and_concentrado(sample_project, tmp_path):
    """El PDF debe existir y tener multiples paginas (portada + concentrado)."""
    out = tmp_path / "diagnostico.pdf"
    pdf_writer.generar_pdf(sample_project, str(out))
    assert out.exists()
    size = out.stat().st_size
    assert size > 1000, f"PDF demasiado pequeno ({size} bytes), probablemente vacio"

    # Magic bytes PDF
    with open(out, "rb") as fh:
        head = fh.read(8)
    assert head.startswith(b"%PDF"), "No es un PDF valido"


# ---------------------------------------------------------------------------
# Tests del orquestador + endpoints
# ---------------------------------------------------------------------------


def test_generar_reporte_masivo_xlsx(tmp_uploads, sample_project):
    """El orquestador debe crear el archivo XLSX en el path esperado."""
    paths = generar_reporte_masivo(sample_project, "proj-1", "rep-1", "XLSX")
    assert paths["fileUrlXlsx"] == "reports/proj-1/rep-1/concentrado.xlsx"
    assert paths["fileUrlPdf"] is None

    full = Path(tmp_uploads) / "reports" / "proj-1" / "rep-1" / "concentrado.xlsx"
    assert full.exists()


def test_generar_reporte_masivo_both(tmp_uploads, sample_project):
    """Formato BOTH genera XLSX + EBOOK (IMPL-20260701-04)."""
    paths = generar_reporte_masivo(sample_project, "proj-2", "rep-2", "BOTH")
    assert paths["fileUrlXlsx"] is not None
    assert paths["fileUrlPdf"] is not None

    base = Path(tmp_uploads) / "reports" / "proj-2" / "rep-2"
    assert (base / "concentrado.xlsx").exists()
    # IMPL-20260701-04: el PDF ahora se nombra EBOOK_{empresa}_{fecha}.pdf.
    ebook_files = list(base.glob("EBOOK_*.pdf"))
    assert len(ebook_files) == 1, (
        f"Esperaba exactamente 1 archivo EBOOK_*.pdf, encontre: "
        f"{[p.name for p in base.glob('*.pdf')]}"
    )
    assert ebook_files[0].stat().st_size > 1000, "EBOOK demasiado pequeno"


def test_generar_reporte_masivo_ebook(tmp_uploads, sample_project):
    """IMPL-20260701-04: format='EBOOK' solo genera el PDF ebook."""
    paths = generar_reporte_masivo(sample_project, "proj-e", "rep-e", "EBOOK")
    assert paths["fileUrlXlsx"] is None
    assert paths["fileUrlPdf"] is not None
    assert paths["fileUrlPdf"].startswith("reports/proj-e/rep-e/EBOOK_")
    assert paths["fileUrlPdf"].endswith(".pdf")

    base = Path(tmp_uploads) / "reports" / "proj-e" / "rep-e"
    assert not (base / "concentrado.xlsx").exists()
    ebook_files = list(base.glob("EBOOK_*.pdf"))
    assert len(ebook_files) == 1
    # Magic bytes PDF
    with open(ebook_files[0], "rb") as fh:
        assert fh.read(8).startswith(b"%PDF")


def test_generar_reporte_masivo_legacy_pdf(tmp_uploads, sample_project):
    """IMPL-20260701-04: 'PDF' legacy sigue funcionando (deprecated, fallback EBOOK)."""
    paths = generar_reporte_masivo(sample_project, "proj-p", "rep-p", "PDF")
    assert paths["fileUrlXlsx"] is None
    assert paths["fileUrlPdf"] is not None
    assert paths["fileUrlPdf"].startswith("reports/proj-p/rep-p/EBOOK_")


def test_generar_reporte_masivo_invalid_format(tmp_uploads, sample_project):
    """Formato invalido debe levantar ValueError."""
    with pytest.raises(ValueError):
        generar_reporte_masivo(sample_project, "p", "r", "TXT")


def test_report_status_transitions_pending_to_ready(tmp_uploads):
    """Job completo: PENDING -> PROCESSING -> READY."""
    # Mock Prisma (AsyncMock para soportar await en BackgroundTask)
    prisma = MagicMock()
    prisma.projectreport = MagicMock()
    prisma.project = MagicMock()
    created = MagicMock(id="rep-success", status="PENDING", projectId="proj-x")
    updated_processing = MagicMock(id="rep-success", status="PROCESSING")
    updated_ready = MagicMock(
        id="rep-success",
        status="READY",
        fileUrlXlsx="reports/proj-x/rep-success/concentrado.xlsx",
        fileUrlPdf=None,
    )
    prisma.projectreport.create = AsyncMock(return_value=created)
    prisma.projectreport.update = AsyncMock(side_effect=[updated_processing, updated_ready])
    prisma.project.find_unique = AsyncMock()

    project = MagicMock(
        id="proj-x",
        company=MagicMock(name="X", legalName=None),
        startDate=datetime(2026, 6, 30),
        workers=[MagicMock(worker=MagicMock(id="w1", firstName="A", lastName="B", universalId="1"), event=None)],
    )
    prisma.project.find_unique.return_value = project

    asyncio.run(
        run_generation_job(
            prisma_client=prisma,
            project_id="proj-x",
            report_id="rep-success",
            format="XLSX",
        )
    )

    # Verificar que se llamo update al menos 2 veces (PROCESSING, READY).
    assert prisma.projectreport.update.call_count == 2
    # Segunda llamada con status=READY.
    final_call_args = prisma.projectreport.update.call_args_list[-1]
    assert final_call_args.kwargs["data"]["status"] == "READY"
    assert final_call_args.kwargs["data"]["fileUrlXlsx"] is not None
    assert (Path(tmp_uploads) / "reports" / "proj-x" / "rep-success" / "concentrado.xlsx").exists()


def test_report_failed_with_error_message(tmp_uploads):
    """Si el proyecto no existe, el job termina con status=FAILED."""
    prisma = MagicMock()
    prisma.projectreport = MagicMock()
    prisma.project = MagicMock()
    prisma.projectreport.update = AsyncMock(return_value=MagicMock())
    prisma.project.find_unique = AsyncMock(return_value=None)  # Proyecto inexistente

    asyncio.run(
        run_generation_job(
            prisma_client=prisma,
            project_id="proj-missing",
            report_id="rep-fail",
            format="XLSX",
        )
    )

    final_call_args = prisma.projectreport.update.call_args_list[-1]
    assert final_call_args.kwargs["data"]["status"] == "FAILED"
    assert "no encontrado" in (final_call_args.kwargs["data"]["errorMessage"] or "").lower()


def test_create_project_report_returns_pending(monkeypatch):
    """El endpoint POST /massive retorna ProjectReport con status PENDING."""
    from fastapi.testclient import TestClient
    from app.api.reports import router as reports_router, set_prisma_client

    prisma = AsyncMock()
    created = MagicMock(
        id="rep-new",
        projectId="proj-1",
        format="XLSX",
        status="PENDING",
        fileUrlXlsx=None,
        fileUrlPdf=None,
        errorMessage=None,
        generatedById="user-1",
        generatedAt=datetime(2026, 6, 30, 12, 0, 0),
        completedAt=None,
    )
    prisma.projectreport.create.return_value = created
    prisma.project.find_unique.return_value = MagicMock(id="proj-1")
    set_prisma_client(prisma)

    from fastapi import FastAPI
    app = FastAPI()
    app.include_router(reports_router)
    client = TestClient(app)

    resp = client.post(
        "/api/v2/projects/proj-1/reports/massive",
        params={"format": "XLSX", "generatedById": "user-1"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "PENDING"
    assert body["projectId"] == "proj-1"
    assert body["format"] == "XLSX"


def test_download_endpoint_rejects_when_not_ready(monkeypatch):
    """Download retorna 409 si status != READY."""
    from fastapi.testclient import TestClient
    from app.api.reports import router as reports_router, set_prisma_client

    prisma = AsyncMock()
    not_ready = MagicMock(
        id="rep-pending",
        projectId="proj-1",
        status="PENDING",
        fileUrlXlsx=None,
        fileUrlPdf=None,
        errorMessage=None,
        generatedById="user-1",
        generatedAt=datetime.utcnow(),
        completedAt=None,
    )
    prisma.projectreport.find_unique.return_value = not_ready
    set_prisma_client(prisma)

    from fastapi import FastAPI
    app = FastAPI()
    app.include_router(reports_router)
    client = TestClient(app)

    resp = client.get(
        "/api/v2/projects/proj-1/reports/rep-pending/download",
        params={"format": "xlsx"},
    )
    assert resp.status_code == 409


def test_api_rejects_legacy_pdf_format(monkeypatch):
    """IMPL-20260701-04: API rechaza 'PDF' (validacion pattern=^(XLSX|EBOOK|BOTH)$)."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.api.reports import router as reports_router, set_prisma_client

    prisma = MagicMock()
    prisma.project.find_unique.return_value = MagicMock(id="proj-1")
    set_prisma_client(prisma)

    app = FastAPI()
    app.include_router(reports_router)
    client = TestClient(app)

    resp = client.post(
        "/api/v2/projects/proj-1/reports/massive",
        params={"format": "PDF", "generatedById": "user-1"},
    )
    # FastAPI responde 422 cuando el parametro no matchea el pattern.
    assert resp.status_code == 422, (
        f"Esperaba 422 (validation error), obtuve {resp.status_code}: {resp.text}"
    )


def test_api_accepts_ebook_format(monkeypatch):
    """IMPL-20260701-04: API acepta 'EBOOK' y crea ProjectReport en PENDING."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.api.reports import router as reports_router, set_prisma_client

    prisma = AsyncMock()
    created = MagicMock(
        id="rep-ebook",
        projectId="proj-1",
        format="EBOOK",
        status="PENDING",
        fileUrlXlsx=None,
        fileUrlPdf=None,
        errorMessage=None,
        generatedById="user-1",
        generatedAt=datetime(2026, 6, 30, 12, 0, 0),
        completedAt=None,
    )
    prisma.projectreport.create.return_value = created
    prisma.project.find_unique.return_value = MagicMock(id="proj-1")
    set_prisma_client(prisma)

    app = FastAPI()
    app.include_router(reports_router)
    client = TestClient(app)

    resp = client.post(
        "/api/v2/projects/proj-1/reports/massive",
        params={"format": "EBOOK", "generatedById": "user-1"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["format"] == "EBOOK"
    assert body["status"] == "PENDING"
    assert body["projectId"] == "proj-1"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])